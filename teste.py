import openpyxl as op
from openpyxl import Workbook, load_workbook

# arquivo = op.load_workbook('dados_pessoais.xlsx') #carrega o arquivo
# plan = arquivo['Dados']  #define a planilha que vai ser usada

# dados = plan['A3'] = 'CPF'
# dados_salvos = arquivo.save('dados_pessoais.xlsx') #salva os dados no arquivo
# # print(dados)

# arquivo = load_workbook('dados_pessoais.xlsx')

# plan = arquivo['Dados']

# dados = plan['A3'].value
# print(dados)

# Cria workbook
wb = op.Workbook()
planilha = wb.active
planilha.title = 'Dados'
planilha.append(['Data', 'Chassi', 'Veículo', 'Cor', 'Combustível', 'Total_Veiculos', 'Destino', 'Hora', 'Revisado', 'Avarias', 'Calib_Pneus', 'Entreg_Téc', 'Litros_Comb'])  # Primeira linha com titulos
wb.save('Veiculos.xlsx')

# wb_carregado = load_workbook('Dados.xlsx')
# sheet_carregado = wb_carregado.active
# print(sheet_carregado)



