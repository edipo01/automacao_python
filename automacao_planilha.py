# import tkinter as tk
# from openpyxl import Workbook
# import os

# def enviar_dados():
#     nome = entry_nome.get().strip()
#     idade = entry_idade.get().strip()
#     cpf = entry_cpf.get().strip()

#     # Adicionar validação para verificar se os campos obrigatórios estão preenchidos
#     if nome and idade and cpf:
#         planilha.append([nome, idade, cpf])
#         wb.save('Dados.xlsx')
#         atualizar_tabela()
#         limpar_campos()
#     else:
#         print("Por favor, preencha todos os campos.")

#     # Código para atualizar a tabela no Excel
#     # Essa função deve ser implementada de acordo com suas necessidades
# def atualizar_tabela():
#     for row in range(1, len(planilha['A']) + 1):
#         for col in range(1, 4):  # Considerando que são 3 colunas
#             cell = planilha.cell(row=row, column=col)
#             if row == 1:  # Cabeçalho
#                 cell.font = Font(bold=True)
#             else:
#                 cell.value = dados[row - 2][col - 1]  # Onde dados é uma lista com os valores a serem atualizados
#     wb.save('Dados.xlsx')

# # Função para limpar os campos
# def limpar_campos():
#     entry_nome.delete(0, 'end')
#     entry_idade.delete(0, 'end')
#     entry_cpf.delete(0, 'end')

# def fechar_programa():
#     root.destroy()

# # Criando a planilha
# wb = Workbook()
# planilha = wb.active
# planilha.title = 'Dados'
# planilha.append(['Nome', 'Idade', 'CPF'])

# # Abrindo a planilha
# os.startfile('Dados.xlsx')

# # Configurações da interface gráfica
# root = tk.Tk()
# root.title("Formulário de Dados")

# label_nome = tk.Label(root, text="Nome:")
# label_nome.pack()
# entry_nome = tk.Entry(root)
# entry_nome.pack()

# label_idade = tk.Label(root, text="Idade:")
# label_idade.pack()
# entry_idade = tk.Entry(root)
# entry_idade.pack()

# label_cpf = tk.Label(root, text="CPF:")
# label_cpf.pack()
# entry_cpf = tk.Entry(root)
# entry_cpf.pack()

# btn_enviar = tk.Button(root, text="Enviar", command=enviar_dados)
# btn_enviar.pack()

# btn_editar = tk.Button(root, text="Editar")  # Botão editar - funcionalidade a adicionar
# btn_editar.pack()

# btn_excluir = tk.Button(root, text="Excluir")  # Botão excluir - funcionalidade a adicionar
# btn_excluir.pack()

# btn_fechar = tk.Button(root, text="Fechar Programa", command=fechar_programa)
# btn_fechar.pack()

# root.mainloop()

# import tkinter as tk
# from openpyxl import Workbook
# import os

# def enviar_dados():
#     nome = entry_nome.get().strip()
#     idade = entry_idade.get().strip()
#     cpf = entry_cpf.get().strip()

#     # Adicionar validação para verificar se os campos obrigatórios estão preenchidos
#     if nome and idade and cpf:
#         planilha.append([nome, idade, cpf])
#         wb.save('Dados.xlsx')
#         atualizar_tabela()
#         limpar_campos()
#     else:
#         print("Por favor, preencha todos os campos.")

#     # Código para atualizar a tabela no Excel
#     # Essa função deve ser implementada de acordo com suas necessidades
# def atualizar_tabela():
#     for row in range(1, len(planilha['A']) + 1):
#         for col in range(1, 4):  # Considerando que são 3 colunas
#             cell = planilha.cell(row=row, column=col)
#             if row == 1:  # Cabeçalho
#                 cell.font = Font(bold=True)
#             else:
#                 cell.value = dados[row - 2][col - 1]  # Onde dados é uma lista com os valores a serem atualizados
#     wb.save('Dados.xlsx')

# # Função para limpar os campos
# def limpar_campos():
#     entry_nome.delete(0, 'end')
#     entry_idade.delete(0, 'end')
#     entry_cpf.delete(0, 'end')

# def fechar_programa():
#     root.destroy()

# # Criando a planilha
# wb = Workbook()
# planilha = wb.active
# planilha.title = 'Dados'
# planilha.append(['Nome', 'Idade', 'CPF'])

# # Abrindo a planilha
# os.startfile('Dados.xlsx')

# # Configurações da interface gráfica
# root = tk.Tk()
# root.title("Formulário de Dados")

# label_nome = tk.Label(root, text="Nome:")
# label_nome.pack()
# entry_nome = tk.Entry(root)
# entry_nome.pack()

# label_idade = tk.Label(root, text="Idade:")
# label_idade.pack()
# entry_idade = tk.Entry(root)
# entry_idade.pack()

# label_cpf = tk.Label(root, text="CPF:")
# label_cpf.pack()
# entry_cpf = tk.Entry(root)
# entry_cpf.pack()

# btn_enviar = tk.Button(root, text="Enviar", command=enviar_dados)
# btn_enviar.pack()

# btn_editar = tk.Button(root, text="Editar")  # Botão editar - funcionalidade a adicionar
# btn_editar.pack()

# btn_excluir = tk.Button(root, text="Excluir")  # Botão excluir - funcionalidade a adicionar
# btn_excluir.pack()

# btn_fechar = tk.Button(root, text="Fechar Programa", command=fechar_programa)
# btn_fechar.pack()

# root.mainloop()


import tkinter as tk
from openpyxl import load_workbook

def enviar_dados():
    nome = entry_nome.get()
    idade = entry_idade.get()
    cpf = entry_cpf.get()

    # Carregar a planilha Excel
    wb = load_workbook('dados_pessoais.xlsx')
    planilha = wb.active

    # Encontrar a próxima linha vazia na planilha
    row = 1
    while planilha.cell(row, 1).value is not None:
        row += 1

    # Atualizar a próxima linha vazia com os novos dados
    planilha.cell(row, 1, value=nome)
    planilha.cell(row, 2, value=idade)
    planilha.cell(row, 3, value=cpf)

    # Salvar a planilha com os novos dados
    wb.save('dados_pessoais.xlsx')

    # Limpar campos de entrada
    entry_nome.delete(0, 'end')
    entry_idade.delete(0, 'end')
    entry_cpf.delete(0, 'end')

# Criar formulário
root = tk.Tk()
root.title("Formulário de Dados")
root.geometry("300x150")

# Campos de entrada
label_nome = tk.Label(root, text="Nome:")
label_nome.pack()
entry_nome = tk.Entry(root)
entry_nome.pack()

label_idade = tk.Label(root, text="Idade:")
label_idade.pack()
entry_idade = tk.Entry(root)
entry_idade.pack()

label_cpf = tk.Label(root, text="CPF:")
label_cpf.pack()
entry_cpf = tk.Entry(root)
entry_cpf.pack()

# Botão de enviar
btn_enviar = tk.Button(root, text="Enviar", command=enviar_dados)
btn_enviar.pack()

# Iniciar o loop principal do formulário
root.mainloop()
