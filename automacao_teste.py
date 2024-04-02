import tkinter as tk
import openpyxl as op

# Criar o workbook e adicionar a primeira linha com títulos
wb = op.Workbook()
planilha = wb.active
planilha.title = 'Dados'
planilha.append(['Data', 'Chassi', 'Veículo', 'Cor', 'Combustível', 'Total_Veiculos', 'Destino', 'Hora', 'Revisado', 'Avarias', 'Calib_Pneus', 'Entreg_Téc', 'Litros_Comb'])
wb.save('Veiculos.xlsx')

def carregar_planilha():
    try:
        wb = op.load_workbook('Veiculos.xlsx')
    except FileNotFoundError:
        wb = op.Workbook()
        planilha = wb.active
        planilha.title = 'Dados'
        planilha.append(['Data', 'Chassi', 'Veículo', 'Cor', 'Combustível'])
        wb.save('Veiculos.xlsx')
        wb = op.load_workbook('Veiculos.xlsx')
    
    return wb

def enviar_dados():
    data = entry_data.get()
    chassi = entry_chassi.get()
    veiculo = entry_veiculo.get()
    cor = entry_cor.get()
    combustivel = entry_combustivel.get()
    total_veiculos = entry_total_veiculos.get()
    destino = entry_destino.get()
    hora = entry_hora.get()
    revisado = entry_revisado.get()
    avarias = entry_avarias.get()
    calib_pneus = entry_calib_pneus.get()
    entreg_tec = entry_entreg_tec.get()
    litros_comb = entry_litros_comb.get()

    # Carregar a planilha Excel
    wb = carregar_planilha()
    planilha = wb.active

    # Encontrar a próxima linha vazia na planilha
    row = 2  # Começa da linha 2 para evitar sobrescrever os títulos
    while planilha.cell(row, 1).value is not None:
        row += 1

    # Atualizar a próxima linha vazia com os novos dados
    planilha.cell(row, 1, value=data)
    planilha.cell(row, 2, value=chassi)
    planilha.cell(row, 3, value=veiculo)
    planilha.cell(row, 4, value=cor)
    planilha.cell(row, 5, value=combustivel)
    planilha.cell(row, 6, value=total_veiculos)
    planilha.cell(row, 7, value=destino)
    planilha.cell(row, 8, value=hora)
    planilha.cell(row, 9, value=revisado)
    planilha.cell(row, 10, value=avarias)
    planilha.cell(row, 11, value=calib_pneus)
    planilha.cell(row, 12, value=entreg_tec)
    planilha.cell(row, 13, value=litros_comb)

    # Salvar a planilha com os novos dados
    wb.save('Veiculos.xlsx')

    # Limpar campos de entrada
    entry_data.delete(0, 'end')
    entry_chassi.delete(0, 'end')
    entry_veiculo.delete(0, 'end')
    entry_cor.delete(0, 'end')
    entry_combustivel.delete(0, 'end')
    entry_total_veiculos.delete(0, 'end')
    entry_destino.delete(0, 'end')
    entry_hora.delete(0, 'end')
    entry_revisado.delete(0, 'end')
    entry_avarias.delete(0, 'end')
    entry_calib_pneus.delete(0, 'end')
    entry_entreg_tec.delete(0, 'end')
    entry_litros_comb.delete(0, 'end')

def buscar_veiculo():
    chassi_desejado = entry_busca_chassi.get()
    wb = carregar_planilha()
    planilha = wb.active
    
    for row in range(2, planilha.max_row + 1):
        if planilha.cell(row, 2).value == chassi_desejado:
            veiculo = planilha.cell(row, 3).value
            resultado_busca.set(f"Chassi: {chassi_desejado}, Veículo: {veiculo}")
            return
    resultado_busca.set("Veículo não encontrado")

# Criar formulário
root = tk.Tk()
root.title("Grandourados Ltda")
root.configure(bg='#008080')  # Cor de fundo do formulário

# Ajustar o layout para preencher toda a tela
root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

# Divisão horizontal com as labels
frame_labels = tk.Frame(root, background="white")
frame_labels.grid(row=0, column=0, sticky="nsew")

label_data = tk.Label(frame_labels, text="Data:", bg='#008080')
label_data.grid(row=0, column=0)

label_chassi = tk.Label(frame_labels, text="Chassi:", bg='#008080')
label_chassi.grid(row=1, column=0)

label_veiculo = tk.Label(frame_labels, text="Veículo:", bg='#008080')
label_veiculo.grid(row=2, column=0)

label_cor = tk.Label(frame_labels, text="Cor:", bg='#008080')
label_cor.grid(row=3, column=0)

label_combustivel = tk.Label(frame_labels, text="Combustível:", bg='#008080')
label_combustivel.grid(row=4, column=0)

label_total_veiculos = tk.Label(frame_labels, text="Total de Veículos:", bg='#008080')
label_total_veiculos.grid(row=5, column=0)

label_destino = tk.Label(frame_labels, text="Destino:", bg='#008080')
label_destino.grid(row=6, column=0)

label_hora = tk.Label(frame_labels, text="Hora:", bg='#008080')
label_hora.grid(row=7, column=0)

label_revisado = tk.Label(frame_labels, text="Revisado:", bg='#008080')
label_revisado.grid(row=8, column=0)

label_avarias = tk.Label(frame_labels, text="Avarias:", bg='#008080')
label_avarias.grid(row=9, column=0)

label_calib_pneus = tk.Label(frame_labels, text="Calibragem dos Pneus:", bg='#008080')
label_calib_pneus.grid(row=10, column=0)

label_entreg_tec = tk.Label(frame_labels, text="Entregue pelo Técnico:", bg='#008080')
label_entreg_tec.grid(row=11, column=0)

label_litros_comb = tk.Label(frame_labels, text="Litros de Combustível:", bg='#008080')
label_litros_comb.grid(row=12, column=0)

# Divisão vertical com os inputs
frame_inputs = tk.Frame(root, bg='#008080')
frame_inputs.grid(row=0, column=1, sticky="nsew")

entry_data = tk.Entry(frame_inputs)
entry_data.grid(row=0, column=0)

entry_chassi = tk.Entry(frame_inputs)
entry_chassi.grid(row=1, column=0)

entry_veiculo = tk.Entry(frame_inputs)
entry_veiculo.grid(row=2, column=0)

entry_cor = tk.Entry(frame_inputs)
entry_cor.grid(row=3, column=0)

entry_combustivel = tk.Entry(frame_inputs)
entry_combustivel.grid(row=4, column=0)

entry_total_veiculos = tk.Entry(frame_inputs)
entry_total_veiculos.grid(row=5, column=0)

entry_destino = tk.Entry(frame_inputs)
entry_destino.grid(row=6, column=0)

entry_hora = tk.Entry(frame_inputs)
entry_hora.grid(row=7, column=0)

entry_revisado = tk.Entry(frame_inputs)
entry_revisado.grid(row=8, column=0)

entry_avarias = tk.Entry(frame_inputs)
entry_avarias.grid(row=9, column=0)

entry_calib_pneus = tk.Entry(frame_inputs)
entry_calib_pneus.grid(row=10, column=0)

entry_entreg_tec = tk.Entry(frame_inputs)
entry_entreg_tec.grid(row=11, column=0)

entry_litros_comb = tk.Entry(frame_inputs)
entry_litros_comb.grid(row=12, column=0)

# Botões e Resultado da Busca
btn_enviar = tk.Button(root, text="Enviar", command=enviar_dados)
btn_enviar.grid(row=1, column=0, columnspan=2)

label_busca_chassi = tk.Label(root, text="Buscar por Chassi:")
label_busca_chassi.grid(row=2, column=0)

entry_busca_chassi = tk.Entry(root)
entry_busca_chassi.grid(row=2, column=1)

btn_buscar_veiculo = tk.Button(root, text="Buscar Veículo", command=buscar_veiculo)
btn_buscar_veiculo.grid(row=3, column=0, columnspan=2)

resultado_busca = tk.StringVar()
resultado_label = tk.Label(root, textvariable=resultado_busca)
resultado_label.grid(row=4, column=0, columnspan=2)

root.mainloop()